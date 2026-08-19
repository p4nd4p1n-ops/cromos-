#!/usr/bin/env python3
"""anade_top50_xlsx.py — añade hoja Top-50 al cromos.xlsx de ayer."""
import json, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX = "/root/.openclaw/workspace/comc/cromos.xlsx"
SRC = "/tmp/scan-top50.json"

rows = json.load(open(SRC))
rows.sort(key=lambda r: (r.get("min") is None, r.get("min") or 0))

wb = load_workbook(XLSX)

# hoja Top-50
if "Top-50" in wb.sheetnames:
    del wb["Top-50"]
ws = wb.create_sheet("Top-50")
headers = ["Jugador", "Mín $", "2º $", "Gap %", "Copias", "Al mín", "Cerca mín",
           "Ventas 7d", "Vel/día", "Días inv.", "Turnover %", "Fecha"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="C00000")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center")

for r in rows:
    ws.append([
        r.get("nombre", ""), r.get("min", ""), r.get("seg", ""), r.get("gap", ""),
        r.get("copias", ""), r.get("n_min", ""), r.get("n_cerca", ""),
        r.get("ventas_7d", ""), r.get("vel_dia", ""), r.get("dias_inv", ""),
        r.get("turnover", ""), r.get("fecha", ""),
    ])

widths = [24, 9, 9, 8, 9, 8, 9, 10, 9, 9, 10, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "$#,##0.00"

# hoja Top-50 Sales
if "Top-50 Sales" in wb.sheetnames:
    del wb["Top-50 Sales"]
ws2 = wb.create_sheet("Top-50 Sales")
ws2.append(["Jugador", "Fecha", "Precio $"])
for r in rows:
    for f, p in r.get("sales", []):
        ws2.append([r["nombre"], f, p])
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 10

wb.save(XLSX)
print(f"OK: {len(rows)} cartas -> {XLSX} hojas={wb.sheetnames}")
