#!/usr/bin/env python3
"""Genera Excel COMC desde el JSON final del día (50 cartas + recent sales)."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "/root/comc-data/scan-2026-08-07.json"
OUT = "/root/comc-data/cromos.xlsx"

rows = json.load(open(SRC))
rows.sort(key=lambda r: (r.get("min") is None, r.get("min") or 0))

wb = Workbook()
ws = wb.active
ws.title = "Cromos"
headers = ["Jugador", "Mín $", "2º $", "Gap %", "Copias", "Al mín", "Cerca mín",
           "Ventas 7d", "Vel/día", "Días inv.", "Turnover %", "Fecha"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="1F4E78")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center")

for r in rows:
    ws.append([
        r.get("nombre", ""), r.get("min", ""), r.get("seg", ""), r.get("gap", ""),
        r.get("copias", ""), r.get("n_min", ""), r.get("n_cerca", ""),
        r.get("ventas_7d", ""), r.get("vel_dia", ""), r.get("dias_inv", ""),
        r.get("turnover", ""), r.get("fecha", ""),
    ])

widths = [24, 9, 9, 8, 9, 8, 9, 10, 9, 9, 10, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "$#,##0.00"

if any(r.get("sales") for r in rows):
    ws2 = wb.create_sheet("Recent Sales")
    ws2.append(["Jugador", "Fecha", "Precio $"])
    for r in rows:
        for f, p in r.get("sales", []):
            ws2.append([r["nombre"], f, p])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 10

wb.save(OUT)
print(f"OK {len(rows)} cartas -> {OUT}")
